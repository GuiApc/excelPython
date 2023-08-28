import openpyxl
from flask import Flask, render_template, request, redirect


app = Flask(__name__)


@app.route('/inserirexcel', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        arquivo_excel = 'Projeto1.xlsx'
        workbook = openpyxl.load_workbook(arquivo_excel)

        tabela_escolhida = request.form['tabela_escolhida']
        if tabela_escolhida == "1":
            planilha = workbook['Planilha1']
        elif tabela_escolhida == "2":
            planilha = workbook['Planilha2']
        elif tabela_escolhida == "3":
            planilha = workbook['Planilha3']
        else:
            return "Escolha inválida. Retorne e tente novamente."

        novos_valores = []
        for frase in ["valor1:", "valor2:", "valor3:", "valor4:"]:
            valor = request.form[frase]
            novos_valores.append(valor)

        linha_insercao = planilha.max_row + 1
        for coluna, valor in enumerate([linha_insercao] + novos_valores, start=1):
            planilha.cell(row=linha_insercao, column=coluna, value=valor)

        workbook.save(arquivo_excel)
        return "Valores inseridos com sucesso!"

    return render_template('index.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)


